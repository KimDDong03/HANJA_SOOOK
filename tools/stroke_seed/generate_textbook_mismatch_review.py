from __future__ import annotations

import argparse
import csv
import html
import re
from collections import defaultdict
from pathlib import Path

import pypdfium2 as pdfium
from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = Path(r"C:\Users\User\Downloads\한자쏘옥완전DB초3~6학년.xlsx")
DEFAULT_MISMATCH_CSV = ROOT / "output/manual_stroke/stroke_mismatch_locations.csv"
DEFAULT_OUT_DIR = ROOT / "output/manual_stroke/textbook_mismatch_review"

TEXTBOOK_PDFS = {
    "초3": Path(r"C:\Users\User\Documents\카카오톡 받은 파일\교과서-3학년_full.pdf"),
    "초4": Path(r"C:\Users\User\Documents\카카오톡 받은 파일\교과서-4학년_full.pdf"),
    "초5": Path(r"C:\Users\User\Documents\카카오톡 받은 파일\교과서-5학년_ful.pdf"),
    "초6": Path(r"C:\Users\User\Documents\카카오톡 받은 파일\교과서-6학년_full.pdf"),
}


def normalize_text(value: object) -> str:
    text = str(value or "")
    text = re.sub(r"[\x00-\x1f\ufffd]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def read_excel_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["📚한자_마스터DB"]
    headers: list[str] | None = None
    rows: list[dict[str, object]] = []

    for row in sheet.iter_rows(values_only=True):
        if row and row[0] == "ID":
            headers = list(row)
            continue
        if headers and row and isinstance(row[0], str) and row[0].startswith("HJ-"):
            item = dict(zip(headers, row))
            if item.get("단원") is not None and item.get("차시") is not None:
                rows.append(item)

    return rows


def lesson_start_page(unit: int, lesson: int) -> int:
    # The supplied textbooks use 26 pages per unit block and 6 pages per lesson.
    return 14 + ((unit - 1) * 26) + ((lesson - 1) * 6)


def learning_page_for_position(unit: int, lesson: int, lesson_position: int) -> int:
    page_offset = 2 if lesson_position <= 4 else 3
    return lesson_start_page(unit, lesson) + page_offset


def render_page(pdf_path: Path, page_number: int, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists():
        return

    document = pdfium.PdfDocument(str(pdf_path))
    try:
        page = document[page_number - 1]
        bitmap = page.render(scale=1.5)
        image = bitmap.to_pil()
        image.save(output_path, quality=88)
    finally:
        document.close()


def extract_snippet(reader: PdfReader, page_number: int, character: str, label: str) -> str:
    text = normalize_text(reader.pages[page_number - 1].extract_text())
    index = text.rfind(label)
    if index < 0:
        index = text.find(character)
    if index < 0:
        return ""
    return text[max(0, index - 90) : index + 260]


def build_review(xlsx_path: Path, mismatch_csv: Path, out_dir: Path) -> None:
    excel_rows = read_excel_rows(xlsx_path)
    rows_by_lesson: dict[tuple[str, int, int], list[dict[str, object]]] = defaultdict(list)
    rows_by_id: dict[str, dict[str, object]] = {}

    for row in excel_rows:
        grade = str(row["학년"])
        unit = int(row["단원"])
        lesson = int(row["차시"])
        rows_by_lesson[(grade, unit, lesson)].append(row)
        rows_by_id[str(row["ID"])] = row

    readers = {grade: PdfReader(str(path)) for grade, path in TEXTBOOK_PDFS.items()}

    with mismatch_csv.open(encoding="utf-8-sig", newline="") as file:
        mismatches = list(csv.DictReader(file))

    review_rows: list[dict[str, object]] = []
    for mismatch in mismatches:
        source = rows_by_id[mismatch["hanja_id"]]
        grade = str(source["학년"])
        unit = int(source["단원"])
        lesson = int(source["차시"])
        lesson_rows = rows_by_lesson[(grade, unit, lesson)]
        lesson_position = next(
            index + 1
            for index, row in enumerate(lesson_rows)
            if row["ID"] == mismatch["hanja_id"]
        )
        page_number = learning_page_for_position(unit, lesson, lesson_position)
        image_name = f"{grade}_{page_number:03d}.jpg"
        render_page(TEXTBOOK_PDFS[grade], page_number, out_dir / "pages" / image_name)

        snippet = extract_snippet(
            readers[grade],
            page_number,
            mismatch["character"],
            normalize_text(source["훈(뜻)"]),
        )
        review_rows.append(
            {
                **mismatch,
                "textbook_page": page_number,
                "lesson_char_position": lesson_position,
                "page_image": f"pages/{image_name}",
                "book_review_count": "",
                "book_review_action": "",
                "book_review_note": "",
                "text_snippet": snippet,
            }
        )

    csv_path = out_dir / "stroke_mismatch_textbook_review.csv"
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=list(review_rows[0].keys()))
        writer.writeheader()
        writer.writerows(review_rows)

    html_path = out_dir / "index.html"
    html_path.write_text(render_html(review_rows), encoding="utf-8")
    print(f"wrote {csv_path}")
    print(f"wrote {html_path}")
    print(f"rows {len(review_rows)}")


def render_html(rows: list[dict[str, object]]) -> str:
    cards = []
    for row in rows:
        title = (
            f"{row['grade']} {row['unit']}-{row['lesson']} "
            f"{row['character']} ({row['meaning']})"
        )
        meta = (
            f"{row['hanja_id']} | {row['unit_name']} | "
            f"현재 {row['current_stroke_count']}획 / CNS {row['cns_stroke_count']}획 | "
            f"교과서 p.{row['textbook_page']} | 차시 {row['lesson_char_position']}번째 글자"
        )
        cards.append(
            f"""
            <article class="card">
              <div class="meta">
                <h2>{html.escape(str(title))}</h2>
                <p>{html.escape(str(meta))}</p>
                <pre>{html.escape(str(row["text_snippet"]))}</pre>
              </div>
              <a href="{html.escape(str(row["page_image"]))}" target="_blank">
                <img src="{html.escape(str(row["page_image"]))}" alt="{html.escape(str(title))}">
              </a>
            </article>
            """
        )

    return f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <title>Textbook Stroke Mismatch Review</title>
  <style>
    body {{
      margin: 0;
      background: #f5f6f8;
      color: #151922;
      font-family: Arial, "Malgun Gothic", sans-serif;
    }}
    header {{
      position: sticky;
      top: 0;
      z-index: 1;
      padding: 16px 24px;
      background: #ffffff;
      border-bottom: 1px solid #d8dde6;
    }}
    h1 {{
      margin: 0 0 4px;
      font-size: 20px;
    }}
    header p {{
      margin: 0;
      color: #5c6678;
      font-size: 13px;
    }}
    main {{
      display: grid;
      gap: 16px;
      padding: 16px;
    }}
    .card {{
      display: grid;
      grid-template-columns: minmax(280px, 380px) minmax(420px, 1fr);
      gap: 16px;
      padding: 16px;
      background: #ffffff;
      border: 1px solid #d8dde6;
      border-radius: 8px;
    }}
    h2 {{
      margin: 0 0 8px;
      font-size: 24px;
    }}
    .meta p {{
      margin: 0 0 12px;
      color: #3d4656;
      line-height: 1.5;
      font-size: 14px;
    }}
    pre {{
      white-space: pre-wrap;
      word-break: keep-all;
      margin: 0;
      padding: 12px;
      max-height: 260px;
      overflow: auto;
      background: #0f172a;
      color: #dbeafe;
      border-radius: 6px;
      font-family: "Malgun Gothic", sans-serif;
      font-size: 13px;
      line-height: 1.5;
    }}
    img {{
      width: 100%;
      max-height: 920px;
      object-fit: contain;
      background: #eef1f5;
      border: 1px solid #d8dde6;
    }}
    @media (max-width: 980px) {{
      .card {{
        grid-template-columns: 1fr;
      }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>교과서 획수 불일치 검토</h1>
    <p>총 {len(rows)}개. CSV의 book_review_count/action/note 칸에 교과서 기준 검토 결과를 적으면 됩니다.</p>
  </header>
  <main>
    {"".join(cards)}
  </main>
</body>
</html>
"""


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--mismatch-csv", type=Path, default=DEFAULT_MISMATCH_CSV)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    args = parser.parse_args()
    build_review(args.xlsx, args.mismatch_csv, args.out_dir)


if __name__ == "__main__":
    main()
