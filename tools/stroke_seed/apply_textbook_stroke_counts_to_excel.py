from __future__ import annotations

import argparse
import csv
import re
import unicodedata
from copy import copy
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = Path(r"C:\Users\User\Downloads\한자쏘옥완전DB초3~6학년.xlsx")
DEFAULT_OUTPUT_XLSX = Path(
    r"C:\Users\User\Downloads\한자쏘옥완전DB초3~6학년_교과서획수반영.xlsx"
)
DEFAULT_REPORT_CSV = ROOT / "output/manual_stroke/textbook_stroke_counts.csv"

TEXTBOOK_PDFS = {
    "초3": Path(r"C:\Users\User\Documents\카카오톡 받은 파일\교과서-3학년_full.pdf"),
    "초4": Path(r"C:\Users\User\Documents\카카오톡 받은 파일\교과서-4학년_full.pdf"),
    "초5": Path(r"C:\Users\User\Documents\카카오톡 받은 파일\교과서-5학년_ful.pdf"),
    "초6": Path(r"C:\Users\User\Documents\카카오톡 받은 파일\교과서-6학년_full.pdf"),
}

HEADER_ROW = 3
ADDED_COLUMNS = ["교과서한자", "교과서획수", "획수출처", "획수검증"]
MANUAL_LESSON_COUNTS = {
    # The textbook PDF prints "yuē" instead of a numeric total count for 約 on
    # 초3 p.152. The stroke-order sequence and the glyph both show 9 strokes.
    ("초3", 6, 2): [5, 7, 9, 7, 13, 12],
}
CJK_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff]")


def normalize_char(value: object) -> str:
    return unicodedata.normalize("NFKC", str(value or "").strip())


def meaning_keyword(value: object) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\([^)]*\)", "", text)
    return text.split()[0] if text.split() else ""


def lesson_start_page(unit: int, lesson: int) -> int:
    # The supplied textbooks use 26 pages per unit block and 6 pages per lesson.
    return 14 + ((unit - 1) * 26) + ((lesson - 1) * 6)


def extract_total_counts(pdf: pdfplumber.PDF, page_number: int, expected_count: int) -> tuple[list[int], str]:
    page = pdf.pages[page_number - 1]
    numbers: list[tuple[float, int]] = []

    for word in page.extract_words(x_tolerance=2, y_tolerance=3):
        text = word["text"]
        if not text.isdecimal() or len(text) > 2:
            continue
        height = float(word["bottom"]) - float(word["top"])
        x = float(word["x0"])
        y = float(word["top"])

        # Total stroke counts are printed as larger numbers in the center label
        # column. Stroke-order labels are tiny and page numbers sit near the foot.
        if height > 8 and 180 < x < 360 and y < 760:
            numbers.append((y, int(text)))

    counts = [count for _, count in sorted(numbers, key=lambda item: item[0])]
    status = "ok" if len(counts) == expected_count else f"expected_{expected_count}_found_{len(counts)}"
    return counts, status


def extract_textbook_labels(
    pdf: pdfplumber.PDF,
    page_number: int,
    expected_count: int,
) -> tuple[list[dict[str, str]], str]:
    page = pdf.pages[page_number - 1]
    labels: list[tuple[float, dict[str, str]]] = []

    for word in page.extract_words(x_tolerance=2, y_tolerance=3):
        text = str(word["text"]).replace("\x00", " ").strip()
        height = float(word["bottom"]) - float(word["top"])
        x = float(word["x0"])
        y = float(word["top"])
        if not (height > 24 and 180 < x < 245 and y < 620):
            continue

        match = CJK_RE.search(text)
        if not match:
            continue
        character = match.group(0)
        tail = text[match.end() :].strip()
        labels.append(
            (
                y,
                {
                    "textbook_character": character,
                    "textbook_character_normalized": normalize_char(character),
                    "textbook_label": tail,
                },
            )
        )

    rows = [label for _, label in sorted(labels, key=lambda item: item[0])]
    status = "ok" if len(rows) == expected_count else f"expected_{expected_count}_labels_found_{len(rows)}"
    return rows, status


def get_header_map(sheet) -> dict[str, int]:
    return {
        str(cell.value): cell.column
        for cell in sheet[HEADER_ROW]
        if cell.value is not None
    }


def read_master_rows(workbook) -> list[dict[str, object]]:
    sheet = workbook["📚한자_마스터DB"]
    headers = get_header_map(sheet)
    rows: list[dict[str, object]] = []
    for row_index in range(HEADER_ROW + 1, sheet.max_row + 1):
        hanja_id = sheet.cell(row_index, headers["ID"]).value
        if not isinstance(hanja_id, str) or not hanja_id.startswith("HJ-"):
            continue
        rows.append(
            {
                "row_index": row_index,
                "id": hanja_id,
                "character": sheet.cell(row_index, headers["한자"]).value,
                "sound": sheet.cell(row_index, headers["음"]).value,
                "meaning": sheet.cell(row_index, headers["훈(뜻)"]).value,
                "grade": sheet.cell(row_index, headers["학년"]).value,
                "unit": int(sheet.cell(row_index, headers["단원"]).value),
                "lesson": int(sheet.cell(row_index, headers["차시"]).value),
                "unit_id": sheet.cell(row_index, headers["단원ID"]).value,
                "unit_name": sheet.cell(row_index, headers["단원명"]).value,
            }
        )
    return rows


def match_textbook_label(
    label: dict[str, str],
    lesson_rows: list[dict[str, object]],
    used_ids: set[str],
) -> tuple[dict[str, object] | None, str]:
    textbook_character = label["textbook_character_normalized"]
    label_keyword = meaning_keyword(label["textbook_label"])

    for row in lesson_rows:
        if str(row["id"]) in used_ids:
            continue
        if normalize_char(row["character"]) == textbook_character:
            return row, "character_match"

    if label_keyword:
        keyword_matches = [
            row
            for row in lesson_rows
            if str(row["id"]) not in used_ids and label_keyword in str(row["meaning"])
        ]
        if len(keyword_matches) == 1:
            return keyword_matches[0], "meaning_match"

    return None, "unmatched_textbook_label"


def build_textbook_counts(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    by_lesson: dict[tuple[str, int, int], list[dict[str, object]]] = {}
    for row in rows:
        key = (str(row["grade"]), int(row["unit"]), int(row["lesson"]))
        by_lesson.setdefault(key, []).append(row)

    pdfs = {grade: pdfplumber.open(path) for grade, path in TEXTBOOK_PDFS.items()}
    try:
        result: list[dict[str, object]] = []
        for key in sorted(by_lesson):
            lesson_rows = by_lesson[key]
            grade, unit, lesson = key
            if len(lesson_rows) != 6:
                raise ValueError(f"{grade} {unit}-{lesson} expected 6 rows, found {len(lesson_rows)}")

            page1 = lesson_start_page(unit, lesson) + 2
            page2 = lesson_start_page(unit, lesson) + 3
            manual_counts = MANUAL_LESSON_COUNTS.get(key)
            labels1, label_status1 = extract_textbook_labels(pdfs[grade], page1, 4)
            labels2, label_status2 = extract_textbook_labels(pdfs[grade], page2, 2)
            labels = labels1 + labels2

            if manual_counts:
                counts = manual_counts
                count_status = "manual_textbook_pdf_typo"
            else:
                counts1, status1 = extract_total_counts(pdfs[grade], page1, 4)
                counts2, status2 = extract_total_counts(pdfs[grade], page2, 2)
                counts = counts1 + counts2
                count_status = "ok" if status1 == "ok" and status2 == "ok" else f"{status1};{status2}"

            label_status = (
                "ok"
                if label_status1 == "ok" and label_status2 == "ok"
                else f"{label_status1};{label_status2}"
            )

            used_ids: set[str] = set()
            for index, label in enumerate(labels):
                row, match_status = match_textbook_label(label, lesson_rows, used_ids)
                if row is not None:
                    used_ids.add(str(row["id"]))
                textbook_page = page1 if index < 4 else page2
                count = counts[index] if index < len(counts) else None
                extraction_status = count_status
                if label_status != "ok":
                    extraction_status = f"{extraction_status};{label_status}"
                if match_status != "character_match":
                    extraction_status = f"{extraction_status};{match_status}"
                result.append(
                    {
                        "id": row["id"] if row else "",
                        "excel_character": row["character"] if row else "",
                        "textbook_character": label["textbook_character_normalized"],
                        "textbook_character_raw": label["textbook_character"],
                        "textbook_label": label["textbook_label"],
                        "sound": row["sound"] if row else "",
                        "meaning": row["meaning"] if row else "",
                        "grade": grade,
                        "unit": unit,
                        "lesson": lesson,
                        "unit_id": row["unit_id"] if row else "",
                        "unit_name": row["unit_name"] if row else "",
                        "lesson_position": index + 1,
                        "textbook_page": textbook_page,
                        "textbook_stroke_count": count if count is not None else "",
                        "character_match_status": match_status,
                        "extraction_status": extraction_status,
                    }
                )
            for row in lesson_rows:
                if str(row["id"]) in used_ids:
                    continue
                result.append(
                    {
                        "id": row["id"],
                        "excel_character": row["character"],
                        "textbook_character": "",
                        "textbook_character_raw": "",
                        "textbook_label": "",
                        "sound": row["sound"],
                        "meaning": row["meaning"],
                        "grade": grade,
                        "unit": unit,
                        "lesson": lesson,
                        "unit_id": row["unit_id"],
                        "unit_name": row["unit_name"],
                        "lesson_position": "",
                        "textbook_page": "",
                        "textbook_stroke_count": "",
                        "character_match_status": "missing_textbook_label",
                        "extraction_status": "missing_textbook_label",
                    }
                )
        return sorted(result, key=lambda item: item["id"])
    finally:
        for pdf in pdfs.values():
            pdf.close()


def ensure_columns(sheet) -> dict[str, int]:
    headers = get_header_map(sheet)
    next_column = sheet.max_column + 1
    for title in ADDED_COLUMNS:
        if title not in headers:
            source = sheet.cell(HEADER_ROW, sheet.max_column)
            target = sheet.cell(HEADER_ROW, next_column)
            target.value = title
            if source.has_style:
                target._style = copy(source._style)
            if source.fill:
                target.fill = copy(source.fill)
            if source.font:
                target.font = copy(source.font)
            if source.alignment:
                target.alignment = copy(source.alignment)
            if source.border:
                target.border = copy(source.border)
            headers[title] = next_column
            next_column += 1
    return headers


def copy_row_style(sheet, source_col: int, target_col: int) -> None:
    for row in range(HEADER_ROW + 1, sheet.max_row + 1):
        source = sheet.cell(row, source_col)
        target = sheet.cell(row, target_col)
        if source.has_style:
            target._style = copy(source._style)
        if source.alignment:
            target.alignment = copy(source.alignment)


def apply_counts_to_workbook(workbook, counts_by_id: dict[str, dict[str, object]]) -> None:
    target_sheets = [
        "📚한자_마스터DB",
        "📖초3학년",
        "📖초4학년",
        "📖초5학년",
        "📖초6학년",
    ]
    for sheet_name in target_sheets:
        sheet = workbook[sheet_name]
        headers = ensure_columns(sheet)
        base_style_col = headers.get("비고") or headers.get("난이도") or sheet.max_column
        for title in ADDED_COLUMNS:
            copy_row_style(sheet, base_style_col, headers[title])

        id_col = headers["ID"]
        count_col = headers["교과서획수"]
        source_col = headers["획수출처"]
        status_col = headers["획수검증"]
        textbook_char_col = headers["교과서한자"]

        for row_index in range(HEADER_ROW + 1, sheet.max_row + 1):
            hanja_id = sheet.cell(row_index, id_col).value
            if hanja_id not in counts_by_id:
                continue
            item = counts_by_id[str(hanja_id)]
            sheet.cell(row_index, textbook_char_col).value = item["textbook_character"]
            sheet.cell(row_index, count_col).value = item["textbook_stroke_count"]
            sheet.cell(row_index, source_col).value = f"교과서 p.{item['textbook_page']}"
            sheet.cell(row_index, status_col).value = item["extraction_status"]

        sheet.column_dimensions[sheet.cell(HEADER_ROW, textbook_char_col).column_letter].width = 12
        sheet.column_dimensions[sheet.cell(HEADER_ROW, count_col).column_letter].width = 12
        sheet.column_dimensions[sheet.cell(HEADER_ROW, source_col).column_letter].width = 16
        sheet.column_dimensions[sheet.cell(HEADER_ROW, status_col).column_letter].width = 18


def write_report(rows: list[dict[str, object]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--output-xlsx", type=Path, default=DEFAULT_OUTPUT_XLSX)
    parser.add_argument("--report-csv", type=Path, default=DEFAULT_REPORT_CSV)
    args = parser.parse_args()

    workbook = load_workbook(args.xlsx)
    master_rows = read_master_rows(workbook)
    report_rows = build_textbook_counts(master_rows)
    write_report(report_rows, args.report_csv)
    apply_counts_to_workbook(workbook, {str(row["id"]): row for row in report_rows})
    workbook.save(args.output_xlsx)

    statuses: dict[str, int] = {}
    for row in report_rows:
        statuses[str(row["extraction_status"])] = statuses.get(str(row["extraction_status"]), 0) + 1
    print(f"wrote {args.report_csv}")
    print(f"wrote {args.output_xlsx}")
    print(f"rows {len(report_rows)}")
    print(f"statuses {statuses}")


if __name__ == "__main__":
    main()
